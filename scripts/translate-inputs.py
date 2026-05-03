"""translate-inputs.py - Convert various input file formats to markdown or text for skill consumption.

Handles: .docx, .doc (via COM on Windows), .xlsx, .csv, .html, .htm, .txt, .json, URLs

Smart .docx handling:
    - If the document contains images: converts to .md with images extracted to a media/ subfolder
    - If the document is text-only: converts to .txt
    - Use --format md or --format txt to override the auto-detection

Usage:
    python scripts/translate-inputs.py <input-folder-or-file> [--output <output-folder>] [--format md|txt|auto]
    python scripts/translate-inputs.py --url <url> [--output <output-folder>]
    python scripts/translate-inputs.py input/specs/wave-planning/

If --output is not specified, converted files are written next to the originals.

Exit codes:
    0 = success (at least one file converted)
    1 = no files found or all failed
"""

import sys
import os
import csv
import json
import re
import argparse
import subprocess
import zipfile
from pathlib import Path


def docx_has_images(filepath):
    """Check if a .docx file contains embedded images.
    .docx files are ZIP archives; images live in word/media/."""
    try:
        with zipfile.ZipFile(filepath, 'r') as zf:
            return any(name.startswith('word/media/') for name in zf.namelist())
    except Exception:
        return False


def is_valid_ooxml(filepath):
    """Check if a file is valid OOXML (.docx) vs an OLE2 .doc with wrong extension."""
    try:
        with zipfile.ZipFile(filepath, 'r') as zf:
            return '[Content_Types].xml' in zf.namelist()
    except Exception:
        return False


def convert_docx_to_md(filepath, extract_media_dir=None):
    """Convert .docx to markdown using pandoc (with optional image extraction) or mammoth."""
    # Try pandoc first
    try:
        cmd = ["pandoc", str(filepath), "-f", "docx", "-t", "markdown", "--wrap=none"]
        if extract_media_dir:
            cmd += ["--extract-media", str(extract_media_dir)]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout
    except Exception:
        pass
    # Fall back to mammoth
    try:
        import mammoth
        with open(filepath, "rb") as f:
            result = mammoth.convert_to_markdown(f)
            return result.value
    except Exception:
        return None


def convert_docx_to_txt(filepath):
    """Convert .docx to plain text using pandoc or python-docx."""
    # Try pandoc plain text output
    try:
        result = subprocess.run(
            ["pandoc", str(filepath), "-f", "docx", "-t", "plain", "--wrap=none"],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout
    except Exception:
        pass
    # Fall back to python-docx paragraph extraction
    try:
        from docx import Document
        doc = Document(filepath)
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception:
        return None


def convert_doc_to_md_via_com(filepath):
    """Convert legacy .doc via Word COM automation (Windows only), then mammoth."""
    try:
        # Convert .doc to .docx first using Word COM via PowerShell
        docx_path = str(filepath) + "x"  # temp .docx
        ps_cmd = f'''
$word = New-Object -ComObject Word.Application
$word.Visible = $false
$doc = $word.Documents.Open("{filepath}")
$doc.SaveAs([ref]"{docx_path}", [ref]16)
$doc.Close()
$word.Quit()
'''
        subprocess.run(["powershell", "-NoProfile", "-Command", ps_cmd],
                       capture_output=True, timeout=30)
        if os.path.exists(docx_path):
            md = convert_docx_to_md(docx_path)
            os.unlink(docx_path)
            return md
        return None
    except Exception:
        return None


def extract_ole2_text(filepath):
    """Extract readable text from an OLE2 .doc file without Word COM.
    Falls back to reading raw bytes and extracting printable text runs."""
    # Try the extract-doc-text.ps1 script if available
    try:
        script_path = Path(__file__).parent / "extract-doc-text.ps1"
        if script_path.exists():
            result = subprocess.run(
                ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass",
                 "-File", str(script_path), "-InputDir", str(filepath.parent)],
                capture_output=True, text=True, timeout=30
            )
            # Check if the script produced an .md file
            md_path = filepath.with_suffix('.md')
            if md_path.exists():
                text = md_path.read_text(encoding='utf-8')
                if text.strip():
                    return text
    except Exception:
        pass
    # Last resort: extract printable text runs from binary
    try:
        with open(filepath, 'rb') as f:
            raw = f.read()
        # Extract UTF-16LE text runs (common in OLE2 .doc)
        text_parts = []
        i = 0
        while i < len(raw) - 1:
            # Look for runs of printable UTF-16LE characters
            char = raw[i:i+2]
            if len(char) == 2 and char[1] == 0 and 32 <= char[0] < 127:
                run = []
                while i < len(raw) - 1 and raw[i+1] == 0 and 32 <= raw[i] < 127:
                    run.append(chr(raw[i]))
                    i += 2
                if len(run) >= 5:  # Only keep runs of 5+ printable chars
                    text_parts.append(''.join(run))
                continue
            i += 1
        if text_parts:
            return '\n'.join(text_parts)
    except Exception:
        pass
    return None


def convert_xlsx_to_md(filepath):
    """Convert Excel workbook to markdown tables."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        md_parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            md_parts.append(f"## {sheet_name}\n")
            # Header row
            headers = [str(c) if c is not None else "" for c in rows[0]]
            md_parts.append("| " + " | ".join(headers) + " |")
            md_parts.append("| " + " | ".join(["---"] * len(headers)) + " |")
            # Data rows
            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                # Pad or trim to match header count
                while len(cells) < len(headers):
                    cells.append("")
                md_parts.append("| " + " | ".join(cells[:len(headers)]) + " |")
            md_parts.append("")
        wb.close()
        return "\n".join(md_parts) if md_parts else None
    except Exception:
        return None


def convert_csv_to_md(filepath):
    """Convert CSV file to markdown table."""
    try:
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return None
        md_parts = []
        headers = rows[0]
        md_parts.append("| " + " | ".join(headers) + " |")
        md_parts.append("| " + " | ".join(["---"] * len(headers)) + " |")
        for row in rows[1:]:
            while len(row) < len(headers):
                row.append("")
            md_parts.append("| " + " | ".join(row[:len(headers)]) + " |")
        return "\n".join(md_parts)
    except Exception:
        return None


def convert_html_to_md(filepath_or_content, is_content=False):
    """Convert HTML to markdown."""
    try:
        from markdownify import markdownify
        if is_content:
            html = filepath_or_content
        else:
            with open(filepath_or_content, "r", encoding="utf-8") as f:
                html = f.read()
        return markdownify(html, heading_style="ATX", strip=["img", "script", "style"])
    except Exception:
        return None


def convert_json_to_md(filepath):
    """Convert JSON to formatted markdown."""
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
            # Array of objects: render as table
            headers = list(data[0].keys())
            md = "| " + " | ".join(headers) + " |\n"
            md += "| " + " | ".join(["---"] * len(headers)) + " |\n"
            for item in data:
                row = [str(item.get(h, "")) for h in headers]
                md += "| " + " | ".join(row) + " |\n"
            return md
        else:
            return "```json\n" + json.dumps(data, indent=2) + "\n```"
    except Exception:
        return None


def convert_txt_to_md(filepath):
    """Read plain text as-is."""
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return f.read()
    except Exception:
        return None


def fetch_url_to_md(url):
    """Fetch a URL and convert to markdown."""
    try:
        import requests
        from markdownify import markdownify
        resp = requests.get(url, timeout=15, headers={"User-Agent": "PM-Skills-Translator/1.0"})
        resp.raise_for_status()
        content_type = resp.headers.get("content-type", "")
        if "html" in content_type:
            md = markdownify(resp.text, heading_style="ATX", strip=["img", "script", "style"])
            return f"# Source: {url}\n\n{md}"
        else:
            return f"# Source: {url}\n\n```\n{resp.text[:50000]}\n```"
    except Exception as e:
        return f"# Source: {url}\n\nFailed to fetch: {e}"


def convert_file(filepath, format_override="auto", output_dir=None):
    """Route a file to the appropriate converter.
    Returns (content_string, method_used, output_extension).
    output_extension is '.md' or '.txt'."""
    ext = filepath.suffix.lower()

    if ext == ".md":
        return convert_txt_to_md(filepath), "already markdown", ".md"
    elif ext == ".docx":
        # Check if this is actually an OLE2 .doc with a .docx extension
        if not is_valid_ooxml(filepath):
            # Route through COM automation like a .doc file
            md = convert_doc_to_md_via_com(filepath)
            if md:
                return md, "word-com (ole2)", ".md"
            # Try raw text extraction as fallback
            text = extract_ole2_text(filepath)
            if text:
                return text, "ole2-text", ".txt"
            return None, "failed (ole2 as .docx)", ".md"
        # Smart .docx handling: detect images to choose format
        has_images = docx_has_images(filepath)
        if format_override == "md" or (format_override == "auto" and has_images):
            # Convert to markdown with images extracted
            media_dir = None
            if has_images:
                media_dir = (output_dir or filepath.parent) / "media"
            content = convert_docx_to_md(filepath, extract_media_dir=media_dir)
            if content:
                method = "pandoc+media" if has_images else "pandoc"
                return content, method, ".md"
            return None, "failed", ".md"
        else:
            # Text-only doc or explicit txt override: convert to plain text
            content = convert_docx_to_txt(filepath)
            if content:
                return content, "pandoc-plain", ".txt"
            # Fall back to markdown if plain text extraction fails
            content = convert_docx_to_md(filepath)
            if content:
                return content, "mammoth-fallback", ".md"
            return None, "failed", ".txt"
    elif ext == ".doc":
        # Legacy format: try COM automation
        md = convert_doc_to_md_via_com(filepath)
        if md:
            return md, "word-com", ".md"
        return None, "failed (legacy .doc)", ".md"
    elif ext in (".xlsx", ".xls"):
        return convert_xlsx_to_md(filepath), "openpyxl", ".md"
    elif ext == ".csv":
        return convert_csv_to_md(filepath), "csv", ".md"
    elif ext in (".html", ".htm"):
        return convert_html_to_md(filepath), "markdownify", ".md"
    elif ext == ".json":
        return convert_json_to_md(filepath), "json", ".md"
    elif ext == ".txt":
        return convert_txt_to_md(filepath), "text", ".txt"
    else:
        return None, f"unsupported format ({ext})", ".md"


def process_sources_manifest(manifest_path, output_dir):
    """Process a sources.md file that may contain URLs."""
    content = manifest_path.read_text(encoding="utf-8")
    urls = re.findall(r'https?://[^\s\)]+', content)
    results = []
    for url in urls:
        safe_name = re.sub(r'[^\w]', '_', url)[:80]
        md = fetch_url_to_md(url)
        if md:
            out_path = output_dir / f"url_{safe_name}.md"
            out_path.write_text(md, encoding="utf-8")
            results.append((url, str(out_path), "fetched"))
    return results


def main():
    parser = argparse.ArgumentParser(description="Translate input files to markdown or text")
    parser.add_argument("path", nargs="?", help="Input folder or file path")
    parser.add_argument("--url", help="Fetch a URL and convert to markdown")
    parser.add_argument("--output", "-o", help="Output directory (default: next to input)")
    parser.add_argument("--format", choices=["md", "txt", "auto"], default="auto",
                        help="Output format for .docx: md (markdown with images), txt (plain text), auto (images=md, text-only=txt)")
    args = parser.parse_args()

    converted = 0
    failed = 0

    if args.url:
        md = fetch_url_to_md(args.url)
        if md:
            out_dir = Path(args.output) if args.output else Path(".")
            out_dir.mkdir(parents=True, exist_ok=True)
            safe_name = re.sub(r'[^\w]', '_', args.url)[:80]
            out_path = out_dir / f"url_{safe_name}.md"
            out_path.write_text(md, encoding="utf-8")
            print(f"  OK [{len(md)} chars]: {args.url} -> {out_path}")
            converted += 1
        else:
            print(f"  FAILED: {args.url}")
            failed += 1
    elif args.path:
        input_path = Path(args.path)
        if input_path.is_file():
            files = [input_path]
            out_dir = Path(args.output) if args.output else input_path.parent
        elif input_path.is_dir():
            files = [f for f in input_path.rglob("*") if f.is_file() and f.suffix.lower() != ".md"]
            # Also process sources.md for URLs
            sources = input_path / "sources.md"
            if sources.exists():
                out_dir = Path(args.output) if args.output else input_path
                url_results = process_sources_manifest(sources, out_dir)
                for url, path, status in url_results:
                    print(f"  OK [url]: {url} -> {path}")
                    converted += 1
            out_dir = Path(args.output) if args.output else input_path
        else:
            print(f"Path not found: {args.path}")
            sys.exit(1)

        out_dir.mkdir(parents=True, exist_ok=True)

        for filepath in files:
            if filepath.name == "sources.md":
                continue  # Already processed above
            if filepath.suffix.lower() == ".gitignore":
                continue

            content, method, out_ext = convert_file(filepath, args.format, out_dir)
            if content:
                out_filename = filepath.stem + out_ext
                out_path = out_dir / out_filename

                # Don't overwrite if already exists and is newer
                if out_path.exists() and out_path.stat().st_mtime > filepath.stat().st_mtime:
                    print(f"  SKIP (up to date): {filepath.name}")
                    continue

                out_path.write_text(content, encoding="utf-8")
                print(f"  OK [{method}]: {filepath.name} -> {out_filename}")
                converted += 1
            else:
                print(f"  FAILED [{method}]: {filepath.name}")
                failed += 1
    else:
        parser.print_help()
        sys.exit(1)

    print(f"\nDone. Converted: {converted}, Failed: {failed}")
    sys.exit(0 if converted > 0 or failed == 0 else 1)

if __name__ == "__main__":
    main()
